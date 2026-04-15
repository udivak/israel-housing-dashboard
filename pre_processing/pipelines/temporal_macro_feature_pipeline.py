"""
temporal_macro_feature_pipeline.py
====================================
כל source מנורמל בנפרד לסכמה אחידה, אחר כך pipeline אחד נקי.
"""

import os
import numpy as np
import pandas as pd
from pymongo import MongoClient
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()
MONGO_URI = os.getenv("MONGODB_URI")
DB_NAME = os.getenv("MONGODB_DB_NAME")
OUTPUT_FILE = os.getenv("TEMPORAL_OUTPUT_XLSX", "temporal_features.xlsx")


# ================================================================
# SCHEMATA — כל שורה מנורמלת מחזירה את המבנה הזה בדיוק
# ================================================================
def empty_row():
    return {
        "transaction_date": None,  # datetime
        "price": None,  # float ₪
        "area_sqm": None,  # float | None = לא קיים במקור
        "rooms": None,  # float
        "floor": None,  # float (מומר מעברית)
        "building_floors": None,  # float
        "year_built": None,  # int
        "city": None,  # str
        "neighborhood": None,  # str
        "street": None,  # str
        "deal_nature": None,  # str
        "project_name": None,  # str
        "is_new_project": 0,  # int 0/1
    }


# ================================================================
# SHARED PARSERS
# ================================================================

FLOOR_MAP = {
    "קרקע": 0,
    "קרקע וראשונה": 0,
    "קרקע ועליית גג": 0,
    "קרקע ושניה": 0,
    "ראשונה": 1,
    "שניה": 2,
    "שלישית": 3,
    "רביעית": 4,
    "חמישית": 5,
    "שישית": 6,
    "שביעית": 7,
    "שמינית": 8,
    "תשיעית": 9,
    "עשירית": 10,
    "אחת עשרה": 11,
    "שתים עשרה": 12,
    "שלוש עשרה": 13,
    "ארבע עשרה": 14,
    "חמש עשרה": 15,
    "שש עשרה": 16,
    "שבע עשרה": 17,
    "שמונה עשרה": 18,
    "תשע עשרה": 19,
    "עשרים": 20,
    "עשרים ואחת": 21,
    "עשרים ושתיים": 22,
    "עשרים ושלוש": 23,
    "עשרים וארבע": 24,
    "עשרים וחמש": 25,
    "עליית גג": -1,
    "מרתף": -1,
    "חניה": -1,
}


def _floor(val) -> float:
    if val is None:
        return np.nan
    s = str(val).strip()
    if s in FLOOR_MAP:
        return float(FLOOR_MAP[s])
    try:
        return float(s)
    except ValueError:
        return np.nan


def _float(val) -> float:
    if val is None:
        return np.nan
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(",", "").strip())
    except ValueError:
        return np.nan


def _int(val):
    f = _float(val)
    return int(f) if not np.isnan(f) else None


def _date_nadlan(s) -> pd.Timestamp:
    return pd.to_datetime(s, errors="coerce")


def _date_odata(s) -> pd.Timestamp:
    return pd.to_datetime(s, errors="coerce")


def _date_odata_fallback(s) -> pd.Timestamp:
    return pd.to_datetime(s, format="%d.%m.%Y", errors="coerce")


def _date_tax(s) -> pd.Timestamp:
    if s and str(s).endswith("Z"):
        s = str(s)[:-1]
    return pd.to_datetime(s, errors="coerce")


# ================================================================
# NORMALIZERS — פונקציה נפרדת לכל source
# ================================================================


def normalize_nadlan_gov(p: dict) -> dict:
    row = empty_row()
    row["transaction_date"] = _date_nadlan(p.get("dealDate"))
    row["price"] = _float(p.get("dealAmount"))
    row["area_sqm"] = _float(p.get("assetArea"))
    row["rooms"] = _float(p.get("roomNum"))
    row["floor"] = _floor(p.get("floor"))
    row["building_floors"] = _float(p.get("buildingFloors"))
    row["year_built"] = _int(p.get("yearBuilt"))
    row["street"] = p.get("address")
    row["neighborhood"] = None
    row["city"] = None
    row["deal_nature"] = p.get("dealNature")
    row["project_name"] = None
    row["is_new_project"] = 0
    return row


def normalize_odata_il_nadlan(p: dict) -> dict:
    row = empty_row()
    dt = p.get("transaction_datetime")
    row["transaction_date"] = _date_odata(dt) if dt else _date_odata_fallback(p.get("transaction_date"))
    row["price"] = _float(p.get("price"))
    row["area_sqm"] = None
    row["rooms"] = _float(p.get("rooms"))
    row["floor"] = _floor(p.get("floor"))
    row["building_floors"] = _float(p.get("building_floors"))
    row["year_built"] = _int(p.get("building_year") or p.get("year_built"))
    row["city"] = p.get("city")
    row["neighborhood"] = None
    row["street"] = p.get("display_address") or p.get("full_address")
    row["deal_nature"] = p.get("property_type") or p.get("DEALNATURE")
    row["project_name"] = p.get("project_name")
    row["is_new_project"] = 1 if p.get("new_project_text") else 0
    return row


def normalize_tax_authority(p: dict) -> dict:
    row = empty_row()
    row["transaction_date"] = _date_tax(p.get("transaction_date"))
    row["price"] = _float(p.get("price"))
    row["area_sqm"] = _float(p.get("area_sqm"))
    row["rooms"] = _float(p.get("rooms"))
    row["floor"] = _floor(p.get("floorNo"))
    row["building_floors"] = None
    row["year_built"] = None
    row["city"] = p.get("settlementNameHeb")
    row["neighborhood"] = p.get("neighborhood")

    street = p.get("streetNameHeb")
    house = p.get("houseNum")
    if street and house:
        row["street"] = f"{street} {house}"
    elif street:
        row["street"] = street
    else:
        row["street"] = None

    row["deal_nature"] = p.get("property_type")
    row["project_name"] = None
    row["is_new_project"] = 0
    return row


NORMALIZERS = {
    "nadlan_gov": normalize_nadlan_gov,
    "odata_il_nadlan": normalize_odata_il_nadlan,
    "tax_authority_nadlan": normalize_tax_authority,
}


# ================================================================
# ECONOMIC CALENDAR — ישראל 1983–2026
# ================================================================

# (copied as-is from original script)
ECONOMIC_CALENDAR = {
    1983: {"cpi_general": 191.0, "prime_rate": 50.0, "gdp_growth": 1.7, "unemployment": 4.5, "usd_ils": 0.04},
    1984: {"cpi_general": 444.9, "prime_rate": 100.0, "gdp_growth": 2.2, "unemployment": 4.6, "usd_ils": 0.28},
    1985: {"cpi_general": 185.2, "prime_rate": 45.0, "gdp_growth": 3.5, "unemployment": 6.7, "usd_ils": 1.18},
    1986: {"cpi_general": 19.7, "prime_rate": 20.0, "gdp_growth": 3.6, "unemployment": 7.0, "usd_ils": 1.49},
    1987: {"cpi_general": 16.1, "prime_rate": 18.0, "gdp_growth": 6.1, "unemployment": 6.1, "usd_ils": 1.60},
    1988: {"cpi_general": 16.2, "prime_rate": 18.0, "gdp_growth": 1.9, "unemployment": 6.4, "usd_ils": 1.60},
    1989: {"cpi_general": 20.7, "prime_rate": 20.0, "gdp_growth": 1.3, "unemployment": 8.9, "usd_ils": 1.92},
    1990: {"cpi_general": 17.6, "prime_rate": 14.0, "gdp_growth": 6.2, "unemployment": 9.6, "usd_ils": 2.02},
    1991: {"cpi_general": 18.0, "prime_rate": 15.0, "gdp_growth": 5.2, "unemployment": 10.6, "usd_ils": 2.28},
    1992: {"cpi_general": 11.9, "prime_rate": 11.0, "gdp_growth": 6.8, "unemployment": 11.2, "usd_ils": 2.46},
    1993: {"cpi_general": 10.9, "prime_rate": 10.0, "gdp_growth": 3.4, "unemployment": 10.0, "usd_ils": 2.83},
    1994: {"cpi_general": 12.3, "prime_rate": 14.5, "gdp_growth": 7.1, "unemployment": 7.8, "usd_ils": 3.01},
    1995: {"cpi_general": 10.0, "prime_rate": 15.5, "gdp_growth": 6.8, "unemployment": 6.9, "usd_ils": 3.01},
    1996: {"cpi_general": 11.3, "prime_rate": 15.5, "gdp_growth": 4.6, "unemployment": 6.7, "usd_ils": 3.23},
    1997: {"cpi_general": 9.0, "prime_rate": 13.5, "gdp_growth": 2.9, "unemployment": 7.7, "usd_ils": 3.45},
    1998: {"cpi_general": 8.6, "prime_rate": 13.5, "gdp_growth": 2.3, "unemployment": 8.6, "usd_ils": 3.80},
    1999: {"cpi_general": 1.3, "prime_rate": 9.0, "gdp_growth": 2.9, "unemployment": 8.9, "usd_ils": 4.14},
    2000: {"cpi_general": 0.0, "prime_rate": 8.5, "gdp_growth": 8.9, "unemployment": 8.8, "usd_ils": 4.08},
    2001: {"cpi_general": 1.4, "prime_rate": 7.5, "gdp_growth": -0.4, "unemployment": 9.4, "usd_ils": 4.21},
    2002: {"cpi_general": 6.5, "prime_rate": 9.0, "gdp_growth": -0.7, "unemployment": 10.3, "usd_ils": 4.74},
    2003: {"cpi_general": -1.9, "prime_rate": 5.0, "gdp_growth": 1.5, "unemployment": 10.7, "usd_ils": 4.55},
    2004: {"cpi_general": 1.2, "prime_rate": 3.75, "gdp_growth": 5.1, "unemployment": 10.4, "usd_ils": 4.48},
    2005: {"cpi_general": 2.4, "prime_rate": 4.5, "gdp_growth": 5.1, "unemployment": 9.0, "usd_ils": 4.49},
    2006: {"cpi_general": 0.0, "prime_rate": 5.0, "gdp_growth": 5.5, "unemployment": 8.4, "usd_ils": 4.46},
    2007: {"cpi_general": 0.5, "prime_rate": 4.0, "gdp_growth": 6.3, "unemployment": 7.3, "usd_ils": 4.11},
    2008: {"cpi_general": 4.6, "prime_rate": 2.5, "gdp_growth": 2.5, "unemployment": 6.1, "usd_ils": 3.59},
    2009: {"cpi_general": 3.9, "prime_rate": 1.0, "gdp_growth": 0.8, "unemployment": 7.6, "usd_ils": 3.93},
    2010: {"cpi_general": 2.7, "prime_rate": 2.5, "gdp_growth": 5.5, "unemployment": 6.6, "usd_ils": 3.73},
    2011: {"cpi_general": 2.2, "prime_rate": 3.25, "gdp_growth": 5.1, "unemployment": 5.6, "usd_ils": 3.58},
    2012: {"cpi_general": 1.6, "prime_rate": 2.25, "gdp_growth": 3.1, "unemployment": 6.9, "usd_ils": 3.86},
    2013: {"cpi_general": 1.8, "prime_rate": 1.5, "gdp_growth": 3.3, "unemployment": 6.2, "usd_ils": 3.61},
    2014: {"cpi_general": 0.5, "prime_rate": 0.25, "gdp_growth": 3.5, "unemployment": 5.9, "usd_ils": 3.58},
    2015: {"cpi_general": -0.6, "prime_rate": 0.1, "gdp_growth": 2.5, "unemployment": 5.3, "usd_ils": 3.89},
    2016: {"cpi_general": -0.5, "prime_rate": 0.1, "gdp_growth": 4.0, "unemployment": 4.8, "usd_ils": 3.84},
    2017: {"cpi_general": 0.2, "prime_rate": 0.1, "gdp_growth": 3.5, "unemployment": 4.2, "usd_ils": 3.60},
    2018: {"cpi_general": 0.8, "prime_rate": 0.25, "gdp_growth": 3.5, "unemployment": 4.0, "usd_ils": 3.74},
    2019: {"cpi_general": 0.6, "prime_rate": 0.25, "gdp_growth": 3.8, "unemployment": 3.8, "usd_ils": 3.56},
    2020: {"cpi_general": -0.7, "prime_rate": 0.1, "gdp_growth": -1.9, "unemployment": 4.3, "usd_ils": 3.44},
    2021: {"cpi_general": 2.8, "prime_rate": 0.1, "gdp_growth": 8.6, "unemployment": 5.0, "usd_ils": 3.23},
    2022: {"cpi_general": 5.3, "prime_rate": 2.75, "gdp_growth": 6.5, "unemployment": 3.8, "usd_ils": 3.52},
    2023: {"cpi_general": 3.3, "prime_rate": 4.75, "gdp_growth": 2.0, "unemployment": 3.5, "usd_ils": 3.69},
    2024: {"cpi_general": 3.5, "prime_rate": 4.5, "gdp_growth": 1.5, "unemployment": 4.5, "usd_ils": 3.75},
    2025: {"cpi_general": 3.2, "prime_rate": 4.25, "gdp_growth": 2.5, "unemployment": 4.8, "usd_ils": 3.65},
    2026: {"cpi_general": 2.8, "prime_rate": 3.75, "gdp_growth": 3.0, "unemployment": 4.5, "usd_ils": 3.55},
}

ELECTION_YEARS = {2013, 2015, 2019, 2020, 2021, 2022}


# ================================================================
# MONGO LOADERS
# ================================================================


def get_db():
    return MongoClient(MONGO_URI)[DB_NAME]


def load_transactions(db) -> pd.DataFrame:
    print("📥 Loading transactions...")
    cursor = db["raw_records"].find(
        {"source_name": {"$ne": "cbs_housing"}},
        {"_id": 1, "source_name": 1, "raw_payload": 1},
    )
    return pd.DataFrame(list(cursor))


def load_cbs(db) -> pd.DataFrame:
    print("📥 Loading CBS monthly housing index (series 40010)...")
    cursor = db["raw_records"].find(
        {
            "source_name": "cbs_housing",
            "raw_payload.series_code": 40010,
            "raw_payload.period_type": "month",
        },
        {"raw_payload": 1},
    )
    rows = []
    for doc in cursor:
        p = doc["raw_payload"]
        if p.get("year") is None or p.get("month") is None:
            continue
        rows.append(
            {
                "year": int(p["year"]),
                "month": int(p["month"]),
                "price_index": float(p["current_base"]),
                "mom_change": float(p["percent_change"]) if p.get("percent_change") is not None else np.nan,
                "annual_change": float(p["percent_change_annual"])
                if p.get("percent_change_annual") is not None
                else np.nan,
            }
        )
    df = pd.DataFrame(rows).sort_values(["year", "month"]).reset_index(drop=True)
    print(f"   → {len(df)} records  ({int(df['year'].min())}–{int(df['year'].max())})")
    return df


# ================================================================
# PIPELINE
# ================================================================


def run():
    db = get_db()
    cbs = load_cbs(db)
    raw = load_transactions(db)

    print("🔧 Normalizing per source...")
    normalized = []
    skipped = 0
    for rec in raw.to_dict("records"):
        source = rec.get("source_name", "")
        normalizer = NORMALIZERS.get(source)
        if normalizer is None:
            skipped += 1
            continue
        row = normalizer(rec.get("raw_payload", {}))
        row["_id"] = str(rec["_id"])
        row["source_name"] = source
        normalized.append(row)

    if skipped:
        print(f"   ⚠️  {skipped} rows skipped (unknown source)")

    df = pd.DataFrame(normalized)

    # time features
    df["transaction_date"] = pd.to_datetime(df["transaction_date"], errors="coerce")
    df["year"] = df["transaction_date"].dt.year.astype("Int64")
    df["quarter"] = df["transaction_date"].dt.quarter.astype("Int64")
    df["month"] = df["transaction_date"].dt.month.astype("Int64")

    # property features
    df["property_age"] = df["year"].astype(float) - pd.to_numeric(df["year_built"], errors="coerce")
    df["is_new_build"] = (df["property_age"] <= 3).astype("Int64")
    df["is_old_build"] = (df["property_age"] >= 30).astype("Int64")
    df["floor_ratio"] = pd.to_numeric(df["floor"], errors="coerce") / pd.to_numeric(
        df["building_floors"], errors="coerce"
    )

    # price features
    df["price_per_sqm"] = df["price"] / df["area_sqm"]

    # CBS merge (year + month)
    print("🔗 Merging CBS monthly index...")
    has_time = df["year"].notna() & df["month"].notna()
    df_valid = df[has_time].copy()
    df_nodate = df[~has_time].copy()

    df_valid["year"] = df_valid["year"].astype(int)
    df_valid["month"] = df_valid["month"].astype(int)
    df_valid = df_valid.merge(cbs, on=["year", "month"], how="left")

    df = pd.concat([df_valid, df_nodate], ignore_index=True)

    # real price
    latest_index = float(cbs["price_index"].max())
    df["real_price_factor"] = latest_index / df["price_index"]
    df["real_price"] = df["price"] * df["real_price_factor"]
    df["real_price_per_sqm"] = df["price_per_sqm"] * df["real_price_factor"]

    # macro enrichment
    print("📊 Enriching with macro features...")
    econ = pd.DataFrame.from_dict(ECONOMIC_CALENDAR, orient="index").rename_axis("year").reset_index()
    econ["year"] = econ["year"].astype(int)
    df = df.merge(econ, on="year", how="left")

    rate_series = econ.set_index("year")["prime_rate"]

    def rate_hike(y):
        if pd.isna(y):
            return np.nan
        y = int(y)
        c, p = rate_series.get(y), rate_series.get(y - 1)
        return int(c > p) if c is not None and p is not None else np.nan

    df["real_interest_rate"] = df["prime_rate"] - df["cpi_general"]
    df["housing_cpi_gap"] = df["annual_change"] - df["cpi_general"]
    df["is_rate_hike_year"] = df["year"].map(rate_hike)
    df["is_post_covid"] = df["year"].isin([2021, 2022]).astype("Int64")
    df["is_election_year"] = df["year"].isin(ELECTION_YEARS).astype("Int64")
    df["is_high_inflation"] = (df["cpi_general"] > 3).astype("Int64")
    df["is_low_rate_era"] = (df["prime_rate"] <= 0.25).astype("Int64")

    # log features
    for src_col, log_col in [
        ("price", "log_price"),
        ("price_per_sqm", "log_price_per_sqm"),
        ("real_price", "log_real_price"),
        ("area_sqm", "log_area_sqm"),
    ]:
        df[log_col] = np.log1p(df[src_col].where(df[src_col].notna()))

    # report
    total = len(df)
    print(f"\n{'='*65}")
    print(f"  PIPELINE REPORT  |  {total:,} rows  |  {len(df.columns)} columns")
    print(f"{'='*65}")

    print("\n📌 Rows per source:")
    for src, cnt in df["source_name"].value_counts().items():
        print(f"   {src:<30} {cnt:>8,}")

    print("\n📌 Transaction year distribution:")
    yr_counts = df["year"].value_counts().sort_index()
    mx = yr_counts.max()
    for yr, cnt in yr_counts.items():
        bar = "█" * (cnt * 40 // mx)
        print(f"   {int(yr)}  {cnt:>6,}  {bar}")

    REPORT_COLS = [
        ("price", "raw"),
        ("area_sqm", "null = odata_il_nadlan has no area"),
        ("price_per_sqm", "derived"),
        ("floor", "Hebrew → int"),
        ("year_built", "parsed"),
        ("transaction_date", "parsed"),
        ("price_index", "CBS merge — null if year<1994 or >2025"),
        ("annual_change", "CBS merge"),
        ("real_price", "CPI-adjusted"),
        ("cpi_general", "macro"),
        ("prime_rate", "macro"),
        ("real_interest_rate", "derived macro"),
        ("housing_cpi_gap", "derived macro"),
        ("log_price", "log feature"),
        ("log_real_price", "log feature"),
    ]
    print("\n📌 Null counts:")
    print(f"   {'column':<24} {'null':>8}  {'%':>6}  note")
    print(f"   {'-'*24} {'-'*8}  {'-'*6}  {'-'*36}")
    for col, note in REPORT_COLS:
        n = int(df[col].isnull().sum())
        pct = 100 * n / total
        flag = " ⚠️" if pct > 30 and "null =" not in note else ""
        print(f"   {col:<24} {n:>8,}  {pct:>5.1f}%  {note}{flag}")

    print("\n📌 CBS coverage:")
    missing = df_valid[df_valid["price_index"].isnull()][["year", "month"]].drop_duplicates()
    if missing.empty:
        print("   ✅ All year/month combos matched")
    else:
        yr_range = f"{int(missing['year'].min())}–{int(missing['year'].max())}"
        print(f"   ⚠️  {len(missing)} unmatched combos  (years {yr_range})")
        print("      → עסקאות מחוץ לטווח CBS 1994–2025")

    # save
    print(f"\n💾 Saving {OUTPUT_FILE}...")
    df.to_excel(OUTPUT_FILE, index=False, sheet_name="transactions")
    _format_excel(OUTPUT_FILE, df)
    print(f"✅ Done  →  {OUTPUT_FILE}  ({total:,} rows × {len(df.columns)} cols)")


# ================================================================
# EXCEL FORMATTING
# ================================================================

GROUPS = {
    "identity": {"cols": ["_id", "source_name"], "fill": "D9E1F2"},
    "transaction": {
        "cols": [
            "transaction_date",
            "price",
            "area_sqm",
            "rooms",
            "floor",
            "building_floors",
            "year_built",
            "neighborhood",
            "city",
            "street",
            "deal_nature",
            "project_name",
            "is_new_project",
        ],
        "fill": "E2EFDA",
    },
    "time": {"cols": ["year", "quarter", "month", "property_age", "is_new_build", "is_old_build", "floor_ratio"], "fill": "FFF2CC"},
    "price": {"cols": ["price_per_sqm"], "fill": "FCE4D6"},
    "cbs": {"cols": ["price_index", "mom_change", "annual_change"], "fill": "DDEBF7"},
    "real_price": {"cols": ["real_price_factor", "real_price", "real_price_per_sqm"], "fill": "F4CCFF"},
    "macro": {"cols": ["cpi_general", "prime_rate", "gdp_growth", "unemployment", "usd_ils"], "fill": "FDE9D9"},
    "derived": {
        "cols": [
            "real_interest_rate",
            "housing_cpi_gap",
            "is_rate_hike_year",
            "is_post_covid",
            "is_election_year",
            "is_high_inflation",
            "is_low_rate_era",
        ],
        "fill": "EDEDED",
    },
    "log": {"cols": ["log_price", "log_price_per_sqm", "log_real_price", "log_area_sqm"], "fill": "D0E8D8"},
}

LEGEND_DESC = {
    "_id": "MongoDB document ID",
    "source_name": "מקור הנתון",
    "transaction_date": "תאריך העסקה",
    "price": "מחיר עסקה (₪)",
    "area_sqm": 'שטח נכס מ\"ר — חסר ב-odata_il_nadlan',
    "rooms": "מספר חדרים",
    "floor": "קומה (מומר מעברית למספר)",
    "building_floors": "קומות בבניין — חסר ב-tax_authority",
    "year_built": "שנת בניה — חסר ב-tax_authority",
    "neighborhood": "שכונה — קיים רק ב-tax_authority",
    "city": "עיר — חסר ב-nadlan_gov (יש רק ID)",
    "street": "כתובת",
    "deal_nature": "סוג עסקה",
    "project_name": "שם פרויקט — קיים רק ב-odata",
    "is_new_project": "1 = דירה מפרויקט חדש (odata בלבד)",
    "year": "שנת העסקה",
    "quarter": "רבעון (1–4)",
    "month": "חודש (1–12)",
    "property_age": "גיל הנכס בעת העסקה (שנים)",
    "is_new_build": "1 = נכס עד 3 שנים",
    "is_old_build": "1 = נכס מעל 30 שנה",
    "floor_ratio": "קומה / קומות בבניין (0–1)",
    "price_per_sqm": 'מחיר לכל מ\"ר',
    "price_index": "אינדקס מחירי דיור CBS חודשי (series 40010, 1994–2025)",
    "mom_change": "שינוי חודש-על-חודש % — CBS",
    "annual_change": "שינוי שנתי % — CBS",
    "real_price_factor": "מקדם הצמדה לאינדקס הנוכחי",
    "real_price": "מחיר ריאלי — מנורמל לאינדקס הנוכחי",
    "real_price_per_sqm": "מחיר ריאלי למ\"ר",
    "cpi_general": "אינפלציה כללית % — CBS",
    "prime_rate": "ריבית פריים % — בנק ישראל",
    "gdp_growth": 'צמיחת תמ\"ג % — CBS / IMF',
    "unemployment": "אחוז אבטלה — CBS",
    "usd_ils": "שער דולר ממוצע שנתי — בנק ישראל",
    "real_interest_rate": "prime_rate – cpi_general",
    "housing_cpi_gap": "annual_change – cpi_general",
    "is_rate_hike_year": "1 = שנת העלאת ריבית לעומת הקודמת",
    "is_post_covid": "1 = 2021/2022",
    "is_election_year": "1 = שנת בחירות",
    "is_high_inflation": "1 = אינפלציה > 3%",
    "is_low_rate_era": "1 = ריבית פריים ≤ 0.25%",
    "log_price": "log(1+price)",
    "log_price_per_sqm": "log(1+price_per_sqm)",
    "log_real_price": "log(1+real_price)",
    "log_area_sqm": "log(1+area_sqm)",
}


def _format_excel(path: str, df: pd.DataFrame):
    wb = load_workbook(path)
    ws = wb["transactions"]
    col_to_fill = {c: g["fill"] for g in GROUPS.values() for c in g["cols"]}
    headers = [cell.value for cell in ws[1]]
    thin = Border(bottom=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="DDDDDD"))

    for ci, name in enumerate(headers, 1):
        cell = ws.cell(1, ci)
        fill = col_to_fill.get(name, "4472C4")
        cell.fill = PatternFill("solid", start_color=fill, end_color=fill)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin

    body = Font(name="Arial", size=9)
    left = Alignment(horizontal="left", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = body
            cell.alignment = left
            cell.border = thin

    wide = {"_id", "transaction_date", "neighborhood", "city", "street", "deal_nature", "project_name"}
    for ci, name in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = 24 if name in wide else 20 if name == "source_name" else 14

    ws.freeze_panes = "A2"
    _add_legend(wb)
    wb.save(path)


def _add_legend(wb):
    if "legend" in wb.sheetnames:
        del wb["legend"]
    ws = wb.create_sheet("legend")
    ws.append(["Group", "Column", "Description"])
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
        cell.alignment = Alignment(horizontal="center")
    for gname, g in GROUPS.items():
        for col in g["cols"]:
            ws.append([gname, col, LEGEND_DESC.get(col, "")])
            r = ws.max_row
            ws.cell(r, 1).fill = PatternFill("solid", start_color=g["fill"], end_color=g["fill"])
            for c in range(1, 4):
                ws.cell(r, c).font = Font(name="Arial", size=9)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 58
    ws.freeze_panes = "A2"


if __name__ == "__main__":
    run()

